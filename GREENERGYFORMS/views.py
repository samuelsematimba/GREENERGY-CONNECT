import string
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from .models import Entry
from .models import Qrcodes
from .models import EmpowerCustomer, EmpowerSale, EmpowerClaim
from django.contrib import messages
from django.utils import timezone
import segno
from io import BytesIO
from django.core.files.base import ContentFile
from .models import Excel
from itertools import zip_longest
from .models import Qrcodes_double
from PIL import Image, ImageDraw, ImageFont
import base64
import csv
from django.http import HttpResponse
from .models import EmpowerCustomer, EmpowerClaim
import random
import string
from datetime import date
from django.shortcuts import render, redirect
from django.contrib import messages

# Create your views here.
def index(request):
    '''the home page to greenergy forms'''
    return render(request,'GREENERGYFORMS/HOMEXT.html')

def GREENERGY_CUSTOMER_FORMS(request):
    """the fuel order page for fuel"""
    return render(request, 'GREENERGYFORMS/GREENERGYCUSTOMERFORMSXT.html')
def save(request):
    """saving the form datat to the database"""
    if request.method== "POST":
        location2=request.POST.get('location2')
        customer_name=request.POST.get('customer_name')
        surname=request.POST.get('surname')
        ninnumber=request.POST.get('ninnumber')
        phonenumber=request.POST.get('phonenumber')
        gender=request.POST.get('gender')
        stove=request.POST.get('stove')
        serialnumber=request.POST.get('serialnumber')
        quantity=request.POST.get('quantity')
        subcidy=request.POST.get('subcidy')
        others=request.POST.get('others')
        date=request.POST.get('date')
        expected=request.POST.get('expected')
        cashreceived=request.POST.get('cashreceived')
        national=request.FILES.get('nationalid')
        warranty=request.FILES.get('warrantycard')
        receipt=request.FILES.get('reciepts')

        Entry.objects.create(
            location2=location2.upper(),
            givenname=customer_name.upper(),
            surname=surname.upper(),
            nin=ninnumber.upper(),
            phonenumber=phonenumber,
            gender=gender.upper(),
            stove=stove,
            serialnumber=serialnumber.upper(),
            quantity=quantity,
            subcidy=subcidy,
            others=others,
            date=date,
            expected=expected,
            cashreceived=cashreceived,
            national=national,
            warranty=warranty,
            receipt=receipt,
        )
        return render(request, 'GREENERGYFORMS/GREENERGYCUSTOMERFORMSXT.html')
    return render(request, 'GREENERGYFORMS/GREENERGYCUSTOMERFORMSXT.html')
def view_data(request):
    """to view the data in a data base"""
    entr=Entry.objects.all().order_by('serialnumber')
    return render(request,'GREENERGYFORMS/view_data.html', {'entries':entr})

def qr_codes(request):
    '''a view to recieve the data from the excel file'''
    return render(request,'GREENERGYFORMS/GENERATEXT.html')
def save2(request):
    '''to save the first numbers of the excel'''
    if request.method=='POST':
        uploaded= request.FILES.get("excel")
        
        if not uploaded:
            return HttpResponse("NO FILE UPLOADED PLEASE UPLOAD A FILE")
        if not uploaded.name.endswith(".xlsx"):
            return HttpResponse("THIS IS NOT AN EXCEL FILE")
        
        sem=[]
        jun=[]
        wb= openpyxl.load_workbook(uploaded)
        single2= wb['singleburners']
        double2=wb['doubleburners']
        for i in range (1,4000):
            numb=single2.cell(row=i,column=1).value
            if numb is None:
                break
            sem.append(numb)
        for s in range (1,4000):
            numb2=double2.cell(row=s,column=1).value
            if numb2 is None:
                break
            jun.append(numb2)
        request.session['sem']= sem 
        request.session['jun']= jun
        paired=zip_longest(sem,jun,fillvalue="")

        #return HttpResponse("f successfully extracted")
        data=Qrcodes.objects.all()
    return render(request, 'GREENERGYFORMS/generate.html', {"data":data, "paired":paired})

def save3(request):
    """this is to allow the user to save to the database"""
    single=request.session.get('sem',[])
    double = request.session.get('jun',[])
    for s in single:
        if s == "":
            break
        Qrcodes.objects.create(
            Singleburner=s
        )

    
    for d in double:
        if d == "":
            break
        Qrcodes_double.objects.create(
            Doubleburner=d
    )
    
    

    return render(request, "GREENERGYFORMS/makeqr.html")

  

#we are going to create a model for the excel files so that we can view what has been extracted
    
def makeqr(request):
    '''view to make the qr codes'''
    #getting all the contents from the data base
    records=Qrcodes.objects.filter(qr_generated=False)
    records2=Qrcodes_double.objects.filter(qr_generated=False)
    single=request.session.get('sem',[])
    double = request.session.get('jun',[])
    qr_list=[]
    qr_list2=[]

    
    for record in records:
    #making the qr code and storing it in a variable

        if record.Singleburner:

            qr_code=segno.make_qr(record.Singleburner,version=2)
            #saving the segno image to the buffer memory
            buffer=BytesIO()
            qr_code.save(buffer,kind='png',border=3,scale=15)
            buffer.seek(0)
            #now we hav to save the image to the record in the data base

            image= Image.open(buffer)
            same=ImageDraw.Draw(image)
            name ='GreenergyClimate'
            name2=record.Singleburner
            font= ImageFont.truetype('arial.ttf',30)
            position=(110,0)
            position2=(160,425)
            same.text(position,name,font=font)
            same.text(position2,name2,font=font)

            final_buffer=BytesIO()
            image.save(final_buffer, format="png")
            final_buffer.seek(0)
            qr_image=base64.b64encode(final_buffer.getvalue()).decode('utf-8')
            qr_list.append(qr_image)

            record.qr_generated=True
            record.qr_codeSB.save(
                f"{record.Singleburner}.png",
                ContentFile(final_buffer.read()),
                save=False
            )
            record.save(update_fields=['qr_codeSB','qr_generated'])
            
    for record2 in records2:
        if record2.Doubleburner:
            qr=segno.make_qr(record2.Doubleburner,version=2)
            buffer2=BytesIO()
            qr.save(buffer2,kind="png",border=3, scale=15)
            buffer2.seek(0)
            
            image2 =Image.open(buffer2)
            draw= ImageDraw.Draw(image2)
            name= 'GreenergyClimate'
            font = ImageFont.truetype('arial.ttf',30)
            name3=record2.Doubleburner
            position=(110,0)
            position2=(160,425)
            draw.text(position,name, font=font )
            draw.text(position2,name3,font=font)

            final_buffer2=BytesIO()
            image2.save(final_buffer2,format="png")
            final_buffer2.seek(0)
            qr_image2= base64.b64encode(final_buffer2.getvalue()).decode("utf-8")
            qr_list2.append(qr_image2)

            record2.qr_generated=True

            record2.qr_codeDB.save(
                f"{record2.Doubleburner}.png",
                ContentFile(final_buffer2.read()),
                save=False
            )
            record2.save()
            

    sema2= Qrcodes_double.objects.all()
    sema= Qrcodes.objects.all()

    return render(request,'GREENERGYFORMS/original_qr_codes_view.html', {"sema":sema, "sema2":sema2, "qr_list":qr_list,"qr_list2":qr_list2})

def show_qr(request):
    """show all the qr codes saved in the data base"""
    if request.method=="POST":
        qr=Qrcodes.objects.all()
    return render(request, 'GREENERGYFORMS/qr_codes_view.html',{"qr":qr})

def empower(request):
    '''to record the empower your community entries'''
    return render(request, 'GREENERGYFORMS/EMPOWERREGISTERXT.html')
def save_for_empower(request):
    ''' to save the information coming in from empower'''
    if request.method=='POST':
        customer_name= request.POST.get('customer_name')
        surname= request.POST.get('surname')
        ninnumber= request.POST.get('ninnumber')
        contact= request.POST.get('contact')
        location= request.POST.get('location')
        customer_name= request.POST.get('customer_name2')
        surname= request.POST.get('surname3')
        ninnumber= request.POST.get('ninnumber4')
        contact= request.POST.get('contact2')


# ─── EMPOWER: Register new customer (Back Office) ───────────────────────────
def empower_register(request):
    """
    Back office registers a new stove buyer.
    If a referral code is provided, creates sale + empower claim for recommender.
    """
    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').upper()
        nin = request.POST.get('nin', '').upper()
        outlet = request.POST.get('outlet', '').upper()
        phone_number = request.POST.get('phone_number', '')
        referral_code_used = request.POST.get('referral_code_used', '').strip().upper()
        stove_number = request.POST.get('stove', '').strip().upper()
        compliance_photo = request.FILES.get('photo')

        # Check if customer already exists by NIN or phone
        if EmpowerCustomer.objects.filter(nin=nin).exists():
            messages.error(request, f"A customer with NIN {nin} is already registered.")
            return render(request, 'GREENERGYFORMS/EMPOWERREGISTERXT.html')
        
        if len(phone_number) != 10 or not phone_number.isdigit():
            messages.error(request, "Please enter a valid 10-digit phone number.")
            return render(request, 'GREENERGYFORMS/EMPOWERREGISTERXT.html')
        
        if len(nin) != 14:
            messages.error(request, "Please enter a valid 14-character NIN.")
            return render(request, 'GREENERGYFORMS/EMPOWERREGISTERXT.html')

        if EmpowerCustomer.objects.filter(phone_number=phone_number).exists():
            messages.error(request, f"A customer with phone {phone_number} is already registered.")
            return render(request, 'GREENERGYFORMS/EMPOWERREGISTERXT.html')

        # Create the new customer — referral code auto-generated
        new_customer = EmpowerCustomer.objects.create(
            full_name=full_name,
            nin=nin,
            phone_number=phone_number,
            stove_number=stove_number,
            outlet=outlet,
            compliance_photo=compliance_photo,
        )

        # If a recommender code was entered
        if referral_code_used:
            try:
                recommender = EmpowerCustomer.objects.get(referral_code=referral_code_used)

                # Prevent self referral
                if recommender.nin == new_customer.nin:
                    messages.warning(request, "A customer cannot refer themselves.")
                else:
                    # Record the sale
                    sale = EmpowerSale.objects.create(
                        buyer=new_customer,
                        recommender=recommender,
                        outlet=request.POST.get('outlet', 'Back Office')
                    )
                    # Create pending empower claim for recommender
                    EmpowerClaim.objects.create(
                        recommender=recommender,
                        sale=sale,
                        status='PENDING'
                    )
                    messages.success(
                        request,
                        f"✅ {new_customer.full_name} registered! "
                        f"Referral code: {new_customer.referral_code}. "
                        f"{recommender.full_name} has earned an Empower gift!"
                    )

            except EmpowerCustomer.DoesNotExist:
                messages.warning(
                    request,
                    f"Referral code '{referral_code_used}' not found. "
                    f"Customer registered without referral."
                )
        else:
            messages.success(
                request,
                f"✅ {new_customer.full_name} registered! "
                f"Their referral code is: {new_customer.referral_code}"
            )

        return render(request, 'GREENERGYFORMS/EMPOWERREGISTERXT.html')

    return render(request, 'GREENERGYFORMS/EMPOWERREGISTERXT.html')

def generate_sales_number():
    """
    Generates a unique sales/receipt number like: EMP-20260505-A3K7
    Format: EMP-YYYYMMDD-XXXX  (4 random uppercase letters/digits)
    """
    today = date.today().strftime('%Y%m%d')
    suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"EMP-{today}-{suffix}"

# ─── EMPOWER: Check and claim gifts (Outlet Manager) ────────────────────────
def empower_claim(request):
 
    # ── Handle "new_search" button on the receipt (clears everything) ──
    if request.method == 'POST' and 'new_search' in request.POST:
        return redirect('GREENERGYFORMS:empower_claim')
 
    # ── Handle CHECKOUT submission ──
    if request.method == 'POST' and request.POST.get('checkout') == '1':
        claim_ids = request.POST.getlist('claim_ids')   # all pending claim IDs passed from template
        qty_raw   = request.POST.get('qty', '1')
        customer_id = request.POST.get('customer_id')
 
        try:
            qty = int(qty_raw)
            customer = EmpowerCustomer.objects.get(id=customer_id)
        except (ValueError, EmpowerCustomer.DoesNotExist):
            messages.error(request, 'Invalid request. Please search again.')
            return redirect('GREENERGYFORMS:empower_claim')
 
        # Safety check: cannot check out more than what exists
        ids_to_claim = claim_ids[:qty]   # take only the qty the outlet requested
 
        if not ids_to_claim:
            messages.error(request, 'No claims selected.')
            return redirect('GREENERGYFORMS:empower_claim')
 
        # Generate a single sales number for this entire checkout batch
        sales_number = generate_sales_number()
 
        # Mark each selected claim as CLAIMED and stamp the sales number
        claimed_count = 0
        for claim_id in ids_to_claim:
            try:
                claim = EmpowerClaim.objects.get(id=claim_id, status='PENDING')
                claim.status = 'CLAIMED'
                claim.sales_number = sales_number   # store on the model (see note below)
                claim.date_claimed = date.today()
                claim.save()
                claimed_count += 1
            except EmpowerClaim.DoesNotExist:
                pass  # skip if already claimed by someone else
 
        # Remaining balance after this checkout
        remaining = EmpowerClaim.objects.filter(
            recommender=customer, status='PENDING'
        ).count()
 
        # Build receipt context to show to the outlet
        checkout_receipt = {
            'sales_number':   sales_number,
            'customer_name':  customer.full_name,
            'qty':            claimed_count,
            'remaining':      remaining,
        }
 
        return render(request, 'GREENERGYFORMS/empower_claim100.html', {
            'checkout_receipt': checkout_receipt,
        })
 
    # ── Handle LOOKUP (phone number search) ──
    if request.method == 'POST' and 'lookup' in request.POST:
        phone = request.POST.get('phone_number', '').strip()
        try:
            customer = EmpowerCustomer.objects.get(phone_number=phone)
            pending_claims = EmpowerClaim.objects.filter(
                recommender=customer, status='PENDING'
            ).order_by('date_created')
 
            return render(request, 'GREENERGYFORMS/empower_claim100.html', {
                'customer':       customer,
                'pending_claims': pending_claims,
                'searched_phone': phone,
            })
        except EmpowerCustomer.DoesNotExist:
            messages.error(request, f'No customer found with phone number {phone}.')
            return render(request, 'GREENERGYFORMS/empower_claim100.html', {
                'searched_phone': phone,
            })
 
    # ── GET — show empty search form ──
    return render(request, 'GREENERGYFORMS/empower_claim100.html', {})



# ─── EMPOWER: View all records (Back Office Overview) ───────────────────────
def empower_records(request):
    """
    Back office can see all customers, referrals and claim statuses.
    """
    all_customers = EmpowerCustomer.objects.all().order_by('-date_registered')
    all_claims = EmpowerClaim.objects.all().order_by('-date_created')
    context = {
        'all_customers': all_customers,
        'all_claims': all_claims,
    }
    return render(request, 'GREENERGYFORMS/empower_records.html', context)

def download_template(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="template.csv"'

    writer = csv.writer(response)
    
    # Write header (match your model fields)
    writer.writerow(['name', 'nin', 'phone_number', 'referral_code', 'stove_number', 'outlet'])

    return response

import csv
from django.shortcuts import render, redirect
from .models import EmpowerCustomer

def upload_csv(request):
    if request.method == 'POST':
        file = request.FILES['csv_file']

        # Ensure it's CSV
        if not file.name.endswith('.csv'):
            return render(request, 'upload.html', {'error': 'Not a CSV file'})

        decoded_file = file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file)

        objects = []

        for row in reader:
            obj = EmpowerCustomer(
                full_name=row['name'],
                nin=row['nin'],
                phone_number=row['phone_number'],
                referral_code=row['referral_code'],
                stove_number=row['stove_number'],
                outlet=row['outlet'],
            )
            objects.append(obj)

        # Bulk insert (VERY FAST)
        EmpowerCustomer.objects.bulk_create(objects)

        return redirect('empower_records')

    return render(request, 'empower_records.html')

def upload(request):
    '''to show the landing page of uploading the data'''
    return render(request, 'GREENERGYFORMS/upload.html')


from django.shortcuts import render
from .models import EmpowerClaim  # adjust to your actual import


def empower_verify(request):
    """
    Back-office view to verify an Empower checkout receipt by its sales number.
    Also allows cross-checking the customer name against what the outlet wrote in the report.
    """

    # GET — show blank search form
    if request.method == 'GET':
        return render(request, 'GREENERGYFORMS/empower_verify.html', {})

    # POST — "Verify Another" button (blank form submit with no sales_number)
    sales_number = request.POST.get('sales_number', '').strip().upper()
    report_name  = request.POST.get('report_name', '').strip()

    if not sales_number:
        return render(request, 'GREENERGYFORMS/empower_verify.html', {})

    # Look up ALL claims that share this sales number
    # (one checkout batch can cover multiple claims, all tagged with the same number)
    claims = EmpowerClaim.objects.filter(
        sales_number=sales_number
    ).select_related(
        'recommender',      # the customer who earned the empower
        'sale__buyer',      # the person who was referred / bought
    ).order_by('date_created')

    return render(request, 'GREENERGYFORMS/empower_verify.html', {
        'claims':          claims,          # queryset — empty if not found
        'searched_number': sales_number,    # so the template can show "not found" with the searched number
        'report_name':     report_name,     # name from the outlet report (for cross-check)
    })


#   path('empower/verify/', views.empower_verify, name='empower_verify'),


